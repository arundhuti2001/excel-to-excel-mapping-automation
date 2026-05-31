from __future__ import annotations

from datetime import datetime, date
from pathlib import Path

import openpyxl


# =========================================================
# PATHS
# =========================================================

ROOT = Path(__file__).resolve().parent

INPUT_DIR = ROOT / "input"
OUTPUT_DIR = ROOT / "output"

OUTPUT_DIR.mkdir(exist_ok=True)

TEMPLATE_FILE = INPUT_DIR / "All_Custom_Indices_upload.xlsx"

FINAL_OUTPUT = OUTPUT_DIR / "April_2026_Output.xlsx"


# =========================================================
# DATE PARSER
# =========================================================

def parse_date(value):

    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    # Excel serial dates
    if isinstance(value, (int, float)):

        try:
            return datetime.fromordinal(
                datetime(1900, 1, 1).toordinal() + int(value) - 2
            ).date()

        except:
            return None

    value = str(value).strip()

    if value == "":
        return None

    formats = [
        "%d-%b-%y",
        "%d-%b-%Y",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d",
    ]

    for fmt in formats:

        try:
            return datetime.strptime(value, fmt).date()

        except:
            pass

    return None


# =========================================================
# FIND EXISTING ROW OR CREATE NEW
# =========================================================

def find_or_create_row(ws, target_date):

    target_date = parse_date(target_date)

    if target_date is None:
        return None

    # CHECK EXISTING ROWS
    for row in range(2, ws.max_row + 1):

        existing_date = parse_date(ws.cell(row, 1).value)

        if existing_date == target_date:
            return row

    # CREATE NEW ROW
    new_row = ws.max_row + 1

    ws.cell(new_row, 1).value = target_date
    ws.cell(new_row, 1).number_format = "DD-MMM-YY"

    return new_row


# =========================================================
# COPY B TO M
# =========================================================

def copy_b_to_m(output_ws):

    print("\n====================================")
    print("COPYING B TO M")
    print("====================================")

    file_path = INPUT_DIR / "Index from V6 Apr 2026.xlsx"

    wb = openpyxl.load_workbook(file_path, data_only=False)

    ws = wb["Sheet2"]

    for row in ws.iter_rows(min_row=2):

        dt = parse_date(row[0].value)

        if dt is None:
            continue

        # ONLY APRIL 2026
        if dt.year != 2026 or dt.month != 4:
            continue

        output_row = find_or_create_row(output_ws, dt)

        if output_row is None:
            continue

        # PRESERVE DATE
        output_ws.cell(output_row, 1).value = row[0].value

        # COPY B:M EXACTLY
        for col in range(2, 14):

            value = ws.cell(row[0].row, col).value

            output_ws.cell(output_row, col).value = value

    print("B TO M COMPLETED")


# =========================================================
# COPY COLUMN N
# =========================================================

def copy_n_column(output_ws):

    print("\n====================================")
    print("COPYING COLUMN N")
    print("====================================")

    file_path = INPUT_DIR / "indices apr26.xlsx"

    wb = openpyxl.load_workbook(file_path, data_only=False)

    ws = wb["russell3000growth"]

    for row in ws.iter_rows(min_row=2):

        dt = parse_date(row[0].value)

        if dt is None:
            continue

        # ONLY APRIL 2026
        if dt.year != 2026 or dt.month != 4:
            continue

        output_row = find_or_create_row(output_ws, dt)

        if output_row is None:
            continue

        # COLUMN N = 14
        output_ws.cell(output_row, 14).value = row[1].value

    print("COLUMN N COMPLETED")


# =========================================================
# COPY COLUMN O
# =========================================================

def copy_o_column(output_ws):

    print("\n====================================")
    print("COPYING COLUMN O")
    print("====================================")

    file_path = INPUT_DIR / "FUSOF & FEGF Adjusted Index Value.xlsx"

    wb = openpyxl.load_workbook(file_path, data_only=False)

    # AUTO DETECT FIRST SHEET
    ws = wb[wb.sheetnames[0]]

    print(f"USING SHEET: {ws.title}")

    for row in ws.iter_rows(min_row=2):

        dt = parse_date(row[0].value)

        if dt is None:
            continue

        # ONLY APRIL 2026
        if dt.year != 2026 or dt.month != 4:
            continue

        output_row = find_or_create_row(output_ws, dt)

        if output_row is None:
            continue

        # COLUMN O = 15
        output_ws.cell(output_row, 15).value = row[1].value

    print("COLUMN O COMPLETED")


# =========================================================
# COPY COLUMN P
# =========================================================

def copy_p_column(output_ws):

    print("\n====================================")
    print("COPYING COLUMN P")
    print("====================================")

    file_path = INPUT_DIR / "indices apr26.xlsx"

    wb = openpyxl.load_workbook(file_path, data_only=False)

    ws = wb["Bse Infra (Adj-Nifty500-FBIF)"]

    for row in ws.iter_rows(min_row=2):

        dt = parse_date(row[0].value)

        if dt is None:
            continue

        # ONLY APRIL 2026
        if dt.year != 2026 or dt.month != 4:
            continue

        output_row = find_or_create_row(output_ws, dt)

        if output_row is None:
            continue

        # COLUMN P = 16
        output_ws.cell(output_row, 16).value = row[1].value

    print("COLUMN P COMPLETED")


# =========================================================
# MAIN
# =========================================================

def main():

    print("\n====================================")
    print("LOADING TEMPLATE")
    print("====================================")

    wb = openpyxl.load_workbook(TEMPLATE_FILE)

    ws = wb["Sheet1"]

    # IMPORTANT:
    # DO NOT DELETE ROWS
    # DO NOT DELETE COLUMNS
    # DO NOT DELETE CELLS
    # PRESERVE EXISTING DATA
    # UPDATE ONLY APRIL 2026

    copy_b_to_m(ws)

    copy_n_column(ws)

    copy_o_column(ws)

    copy_p_column(ws)

    print("\n====================================")
    print("SAVING OUTPUT")
    print("====================================")

    wb.save(FINAL_OUTPUT)

    print("\nDONE")
    print(f"\nOUTPUT FILE:\n{FINAL_OUTPUT}")


# =========================================================
# RUN
# =========================================================

if __name__ == "__main__":
    main()